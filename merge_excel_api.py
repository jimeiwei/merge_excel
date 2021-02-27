"""Created by jimw"""
"""Current api file made for merging excel file, point out difference between two different excels"""

import openpyxl
import merge_excel_comm as merge_comm
from merge_excel_comm import TT_ERR
from merge_excel_comm import TT_OK
from merge_excel_comm import EXCEPTION_HAPPEN_FLAG


min_row_excel = 1
min_column_excel = 1

EXCEL_MERGED_CELLS = ''

"""
python merge_excel_run.py /Users/jimeiwei/Desktop/电气部分.xlsx /Users/jimeiwei/Desktop/电气部分.xlsx
python merge_excel_run.py /Users/jimeiwei/Desktop/电气部分.xlsx /Users/jimeiwei/Desktop/电气部分.xlsx

功能点：
1.单元格对比
2.检查文件
3.图形化
"""


"""
打开并返回excel文件词典，如果失败返回数字 TT_ERR
@:parameter     p_filename
@:remake        如果失败返回数字 TT_ERR
@:return        返回excel词典
@:exception     IOError
"""
def merge_excel_load_one_excel(p_filename):
    merge_comm.comm_check_type_str(p_filename)
    if (".xlsx" not in p_filename):
        print("Current app open xlsx file only\n")
        return TT_ERR
    try:
        wb = openpyxl.load_workbook(p_filename, data_only = True)
    except (IOError, FileNotFoundError, KeyError) as e:
        EXCEPTION_HAPPEN_FLAG = 1
        merge_comm.comm_check_err(e)
        return TT_ERR
    else:
        return wb


"""
打开并返回excel的工作薄名称集合，如果失败返回数字 TT_ERR
@:parameter     workbook
@:remake        如果失败返回数字 TT_ERR
@:return        返回excel的工作薄名称集合
@:exception     AttributeError
"""
def merge_excel_load_sheetnames_by_wb(workbook):
    try:
        sheet_names = workbook.sheetnames
    except AttributeError as e:
        EXCEPTION_HAPPEN_FLAG = 1
        merge_comm.comm_check_err(e)
        return TT_ERR
    else:
        return sheet_names


"""
打开某一个sheet页的内容
@:parameter     workbook
@:remake        如果失败返回数字 TT_ERR
@:return        返回excel的工作薄内容
@:exception     TypeError
"""
def merge_excel_load_one_sheet(workbook, sheet_name):
    merge_comm.comm_check_type_str(sheet_name)
    try:
        sheet_content = workbook[sheet_name]
    except (TypeError, AttributeError) as e:
        EXCEPTION_HAPPEN_FLAG = 1
        merge_comm.comm_check_err(e)
        return TT_ERR
    else:
        return sheet_content


"""
获取最大行数
@:parameter     sheet_comtent
@:remake        如果失败返回数字 TT_ERR
@:return        获取最大行数
@:exception     AttributeError
"""
def merge_excel_max_row_get(sheet_content):
    try:
        max_row_currr_sheet = sheet_content.max_row
    except AttributeError as e:
        EXCEPTION_HAPPEN_FLAG = 1
        merge_comm.comm_check_err(e)
        return TT_ERR
    else:
        return max_row_currr_sheet


"""
获取最大列数
@:parameter     sheet_comtent
@:remake        如果失败返回数字 TT_ERR
@:return        获取最大列数
@:exception     AttributeError
"""
def merge_excel_max_column_get(sheet_content):
    try:
        max_cloumn_currr_sheet = sheet_content.max_column
    except AttributeError as e:
        EXCEPTION_HAPPEN_FLAG = 1
        merge_comm.comm_check_err(e)
        return TT_ERR
    else:
        return max_cloumn_currr_sheet


"""
通过wb，sheet页名称，获取最大行数
@:parameter     workbook
@:parameter     sheet_name
@:remake        如果失败返回数字 TT_ERR
@:return        获取最大行数
@:exception     
"""
def merge_excel_max_row_get_by_workbook(workbook, sheet_name):
    sheet_content = merge_excel_load_one_sheet(workbook, sheet_name)
    merge_comm.comm_check_rc(sheet_content)

    max_row = merge_excel_max_row_get(sheet_content)
    merge_comm.comm_check_rc(max_row)

    return TT_OK

"""
通过wb，sheet页名称，获取最大列数
@:parameter     workbook
@:parameter     sheet_name
@:remake        如果失败返回数字 TT_ERR
@:return        获取最大行数
@:exception     AttributeError
"""
def merge_excel_max_column_get_by_workbook(workbook, sheet_name):

    sheet_content = merge_excel_load_one_sheet(workbook, sheet_name)
    merge_comm.comm_check_rc(sheet_content)

    max_column = merge_excel_max_column_get(sheet_content)
    merge_comm.comm_check_rc(max_column)

    return TT_OK



"""
通过行、列索引获取该单元格的内容
@:parameter     sheet_content
@:parameter     row
@:parameter     column
@:remake        如果失败返回数字 TT_ERR
@:return        当前行、列的值
@:exception     AttributeError
"""
def merge_excel_value_get_by_row_column(sheet_content, row, column):
    max_row = 0
    max_column = 0
    rtn = TT_OK

    merge_comm.comm_check_type_int(row)
    merge_comm.comm_check_type_int(column)

    max_row = merge_excel_max_row_get(sheet_content)
    merge_comm.comm_check_rc(max_row)
    max_column = merge_excel_max_column_get(sheet_content)
    merge_comm.comm_check_rc(max_column)

    merge_comm.comm_check_index(row, 0, max_row)
    merge_comm.comm_check_index(column, 0, max_column)

    try:
        cell_value = sheet_content.cell(row = row, column = column).value
    except (AttributeError, ValueError) as e:
        EXCEPTION_HAPPEN_FLAG = 1
        merge_comm.comm_check_err(e)
        return TT_ERR
    else:
        return cell_value


"""
检查两个文件中一个单元格的内容是否正确
@:parameter     call_a_value
@:parameter     call_b_value
@:remake        如果失败返回数字 TT_ERR
@:return        如果正确返回数字 TT_OK
@:exception     
"""
def merge_excel_compatr_two_cell(call_a_value, call_b_value):
    merge_comm.comm_check_type_str(call_a_value)
    merge_comm.comm_check_type_str(call_b_value)

    if (call_a_value != call_b_value):
        return TT_ERR
    else:
        return TT_OK

"""
将两个文件中相同行、列单元格merge过去，默认是从call_a_value移动到call_b_value
@:parameter     sheet_content_a
@:parameter     sheet_content_b
@:parameter     row
@:parameter     column
@:parameter     order, 0：call_a_value移动到call_b_value， 1：call_b_value移动到call_a_value
@:remake        该函数执行完未保存文件，需要在上层接口中保存
@:return        如果失败返回数字 TT_ERR,如果正确返回数字 TT_OK
@:exception  
"""
def merge_excel_merge_one_cell(sheet_content_a, sheet_content_b, row, column, order = 0):
    value_a = ''
    value_b = ''

    merge_comm.comm_check_type_int(row)
    merge_comm.comm_check_type_int(column)
    merge_comm.comm_check_type_int(order)

    value_a = merge_excel_value_get_by_row_column(sheet_content_a, row, column)
    merge_comm.comm_check_rc(value_a)
    value_b = merge_excel_value_get_by_row_column(sheet_content_b, row, column)
    merge_comm.comm_check_rc(value_b)

    if (value_a == value_b):
        return TT_OK
    else:
        if (order == 0):
            sheet_content_a.cell(row = row, column = column).value = value_b
        else:
            sheet_content_b.cell(row = row, column = column).value = value_a
    return TT_OK



"""
merge两个文件中的一整行，默认是从call_a_value移动到call_b_value (多个单元格)
@:parameter     sheet_content_a
@:parameter     sheet_content_b
@:parameter     row
@:parameter     order, 0：call_a_value移动到call_b_value， 1：call_b_value移动到call_a_value
@:remake        该函数执行完未保存文件，需要在上层接口中保存
@:return        如果失败返回数字 TT_ERR,如果正确返回数字 TT_OK
@:exception  
"""
def merge_excel_merge_one_row(sheet_content_a, sheet_content_b, row, order = 0):
    rtn = TT_OK
    max_column_excel_a = 0
    max_column_excel_b = 0
    max_column = max_column_excel_a
    merge_comm.comm_check_type_int(row)
    merge_comm.comm_check_type_int(order)

    max_column_excel_a = merge_excel_max_column_get(sheet_content_a)
    merge_comm.comm_check_rc(max_column_excel_a)

    max_column_excel_b = merge_excel_max_column_get(sheet_content_b)
    merge_comm.comm_check_rc(max_column_excel_b)

    if(max_column_excel_b >=  max_column_excel_a):
        max_column = max_column_excel_b
    cloumn_list = list(range(0, max_column))

    for column in cloumn_list:
        rtn = merge_excel_merge_one_cell(sheet_content_a, sheet_content_b, row, column, order)
        merge_comm.comm_check_rc(rtn)

    return TT_OK


"""
merge两个文件中的一整列，默认是从call_a_value移动到call_b_value (多个单元格)
@:parameter     sheet_content_a
@:parameter     sheet_content_b
@:parameter     column
@:parameter     order, 0：call_a_value移动到call_b_value， 1：call_b_value移动到call_a_value
@:remake        该函数执行完未保存文件，需要在上层接口中保存
@:return        如果失败返回数字 TT_ERR,如果正确返回数字 TT_OK
@:exception  
"""
def merge_excel_merge_one_column(sheet_content_a, sheet_content_b, column, order=0):
    rtn = TT_OK
    max_row_excel_a = 0
    max_row_excel_b = 0
    max_row = max_row_excel_a
    merge_comm.comm_check_type_int(column)
    merge_comm.comm_check_type_int(order)

    max_row_excel_a = merge_excel_max_row_get(sheet_content_a)
    max_row_excel_b += merge_excel_max_row_get(sheet_content_b)
    merge_comm.comm_check_rc(rtn)

    if (max_row_excel_b >= max_row_excel_a):
        max_row = max_row_excel_b
    row_list = list(range(0, max_row))

    for row in row_list:
        rtn = merge_excel_merge_one_cell(sheet_content_a, sheet_content_b, row, column, order)
        merge_comm.comm_check_rc(rtn)

    return TT_OK


"""
保存文件
@:parameter     filename
@:remake        
@:return        如果失败返回数字 TT_ERR,如果正确返回数字 TT_OK
@:exception  
"""
def merge_excel_save_file(filename):

    wb = merge_excel_load_one_excel(filename)
    merge_comm.comm_check_rc(wb)

    wb.save(filename)

    return TT_OK


"""
检查两个sheet内容是否有差异，返回差异点
@:parameter     p_filename
@:remake        如果失败返回数字 TT_ERR
@:return        存储差异点的词典
@:exception     IOError
"""
def merge_excel_check_sheet_diff(sheet_content_a, sheet_content_b):
    value_cell_a = ''
    value_cell_b = ''

    max_row_excel_a = merge_excel_max_row_get(sheet_content_a)
    merge_comm.comm_check_rc(max_row_excel_a)

    max_row_excel_b = merge_excel_max_row_get(sheet_content_b)
    merge_comm.comm_check_rc(max_row_excel_b)

    max_column_excel_a = merge_excel_max_column_get(sheet_content_a)
    merge_comm.comm_check_rc(max_column_excel_a)

    max_column_excel_b = merge_excel_max_column_get(sheet_content_b)
    merge_comm.comm_check_rc(max_column_excel_b)

    for row in range(min_row_excel, max([max_row_excel_a, max_row_excel_b])):
        for column in range(min_column_excel, max([max_column_excel_a, max_column_excel_b])):
            value_cell_a = merge_excel_value_get_by_row_column(sheet_content_a, row, column)
            merge_comm.comm_check_rc(value_cell_a)

            value_cell_b = merge_excel_value_get_by_row_column(sheet_content_b, row, column)
            merge_comm.comm_check_rc(value_cell_b)

            if value_cell_a != value_cell_b:
                return TT_ERR

    return TT_OK

"""
检查两个文件内容是否有差异，返回有差异点的sheet页
@:parameter     p_filename
@:remake        如果失败返回数字 TT_ERR
@:return        存储差异点的词典
@:exception     IOError
"""
def merge_excel_check_workbook_diff(excel_file_a, excel_file_b):
    rtn = TT_OK

    diff_sheets = []

    workbook_a = ''
    sheet_names_a = []
    sheet_content_a = ''
    cell_value_a = ''

    workbook_b = ''
    sheet_names_b = []
    sheet_content_b = ''
    cell_value_b = ''

    merge_comm.comm_check_type_str(excel_file_a)
    merge_comm.comm_check_type_str(excel_file_b)

    workbook_a = merge_excel_load_one_excel(excel_file_a)
    merge_comm.comm_check_rc(workbook_a)
    sheet_names_a = merge_excel_load_sheetnames_by_wb(workbook_a)
    merge_comm.comm_check_rc(sheet_names_a)

    workbook_b = merge_excel_load_one_excel(excel_file_b)
    merge_comm.comm_check_rc(workbook_b)
    sheet_names_b = merge_excel_load_sheetnames_by_wb(workbook_b)
    merge_comm.comm_check_rc(sheet_names_b)

    for steet_name_a, steet_name_b in zip(sheet_names_a, sheet_names_b):
        if steet_name_a != steet_name_b:
            diff_sheets.append(sheet_names_a.index(steet_name_a))
        else:
            sheet_content_a = merge_excel_load_one_sheet(workbook_a, steet_name_a)
            merge_comm.comm_check_rc(sheet_content_a)

            sheet_content_b = merge_excel_load_one_sheet(workbook_b, steet_name_b)
            merge_comm.comm_check_rc(sheet_content_b)

            rtn = merge_excel_check_sheet_diff(sheet_content_a, sheet_content_b)

            if rtn != TT_OK:
                diff_sheets.append(sheet_names_a.index(steet_name_a))

    return diff_sheets


"""获取合并单元格列表"""
def merge_excel_merge_cells_get(sheet_content):
    try:
        excel_merge_cells = sheet_content.merge_cells.ranges
    except AttributeError as e:
        EXCEPTION_HAPPEN_FLAG = 1
        merge_comm.comm_check_err(e)
        return TT_ERR
    else:
        return excel_merge_cells


"""检查某个元格是否是合并单元格"""
def merge_excel_merge_cells_check(sheet_content, row, col):
    merged_cells = merge_excel_merge_cells_get(sheet_content)

    for i in merged_cells:
        if (row, col) in merged_cells[i].top:
            return [1, i]
    return [0, 0]


"""再确定是单元格的情况下获取合并单元格的范围"""
def merge_excel_merged_cell_ranges_get(sheet_content, row, col):
    if(merge_excel_merge_cells_check(sheet_content, row, col)[1]):
        return merge_excel_merge_cells_get(sheet_content)[merge_excel_merge_cells_check(sheet_content, row, col)[1]].top
    else:
        return TT_ERR
